from openpyxl import Workbook, load_workbook

fn = r'f:\test.xlsx'				# 文件名
wb = Workbook()				        # 创建工作簿
ws = wb.create_sheet(title='你好，世界')	        # 创建工作表
ws['A1'] = '这是第一个单元格'			# 单元格赋值
ws['B1'] = 3.1415926
wb.save(fn)					# 保存Excel文件
wb = load_workbook(fn)	   		        # 打开已有的Excel文件
ws = wb.worksheets[1]				# 打开指定索引的工作表
print(ws['A1'].value)				# 读取并输出指定单元格的值
ws.append([1,2,3,4,5])			        # 添加一行数据
ws.merge_cells('F2:F3')			        # 合并单元格
ws['F2'] = '=sum(A2:E2)'			# 写入公式
for r in range(10, 15):
    for c in range(3, 8):
        ws.cell(row=r, column=c, value=r*c)	# 写入单元格数据
wb.save(fn)
